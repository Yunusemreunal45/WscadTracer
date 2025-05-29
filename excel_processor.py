import pandas as pd
import numpy as np
import os
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime
import json
import glob

class ExcelProcessor:
    """Class for processing and comparing WSCAD BOM Excel files"""

    def __init__(self):
        """Initialize Excel processor with WSCAD specific configurations"""
        self.default_sheet_name = 'Sayfa1'
        self.header_row = 6  # Based on analysis: headers are in row 6 (1-based)
        self.data_start_row = 7  # Data starts from row 7 (1-based)
        
        # WSCAD specific column mappings (Turkish)
        self.wscad_columns = {
            'POZ_NO': 'POZ NO',
            'PARCA_NO': 'PARCA NO', 
            'PARCA_ADI': 'PARCA ADI',
            'BIRIM_ADET': 'BİRİM\nADET',
            'TOPLAM_ADET': 'TOPLAM\nADET',
            'AGIRLIK': 'AGIRLIK',
            'TANIM': 'TANIM',
            'BOY_KALINLIK': 'BOY/KALINLIK',
            'MALZEME': 'MALZEME',
            'ACIKLAMA': 'ACIKLAMA',
            'NOT': 'NOT',
            'K_YERI': 'K. YERI',
            'STOK_KODU': 'STOK KODU',
            'REV_NO': 'REV. NO',
            'METOT_KONTROL': 'METOT KONTROL  AÇIKLAMASI'
        }
        
        # Critical columns for BOM comparison (most important changes)
        self.critical_columns = ['POZ NO', 'PARCA NO', 'PARCA ADI', 'TOPLAM\nADET', 'STOK KODU']
        
        # Define color fills for changed cells
        self.added_fill = PatternFill(start_color='87CEFA', end_color='87CEFA', fill_type='solid')  # Light blue
        self.removed_fill = PatternFill(start_color='FF6347', end_color='FF6347', fill_type='solid')  # Tomato red
        self.changed_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')  # Light yellow
        self.quantity_increase_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green
        self.quantity_decrease_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # Orange
        self.zero_quantity_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red

    def is_wscad_excel(self, filepath):
        """Check if file is a WSCAD Excel file"""
        if not filepath.lower().endswith(('.xlsx', '.xls')):
            return False
            
        try:
            # Quick check for WSCAD structure
            df = pd.read_excel(filepath, sheet_name=self.default_sheet_name, header=None, nrows=10)
            
            # Look for WSCAD indicators in the first few rows
            wscad_indicators = ['İŞ EMRİ NO', 'PARCA NO', 'PARCA ADI', 'TOPLAM\nADET', 'POZ NO']
            file_content = df.to_string().upper()
            
            indicator_count = sum(1 for indicator in wscad_indicators if indicator.upper() in file_content)
            return indicator_count >= 3  # At least 3 indicators should be present
            
        except Exception:
            return False

    def process_file(self, filepath):
        """Process a WSCAD Excel file for later comparison"""
        try:
            # Check if file exists
            if not os.path.exists(filepath):
                raise FileNotFoundError(f"File not found: {filepath}")

            # Check if this is a WSCAD Excel file
            if not self.is_wscad_excel(filepath):
                print(f"Warning: {filepath} may not be a WSCAD BOM Excel file")

            # Load the Excel file with proper header detection
            df = pd.read_excel(filepath, sheet_name=self.default_sheet_name, header=self.header_row-1)
            
            # Clean column names (remove line breaks and extra spaces)
            df.columns = [str(col).strip().replace('\r\n', '\n') for col in df.columns]
            
            # Get file metadata
            excel_file = pd.ExcelFile(filepath)
            
            # Extract project information from the header area
            header_df = pd.read_excel(filepath, sheet_name=self.default_sheet_name, header=None, nrows=6)
            project_info = self._extract_project_info(header_df)
            
            return {
                'filepath': filepath,
                'filename': os.path.basename(filepath),
                'sheet_count': len(excel_file.sheet_names),
                'row_count': len(df),
                'column_count': len(df.columns),
                'processed_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'project_info': project_info,
                'columns': list(df.columns)
            }
        except Exception as e:
            raise Exception(f"Error processing WSCAD file: {e}")

    def _extract_project_info(self, header_df):
        """Extract project information from header rows"""
        project_info = {}
        try:
            # Convert header area to string for easier searching
            for idx, row in header_df.iterrows():
                row_str = ' '.join([str(cell) for cell in row if pd.notna(cell)])
                
                # Extract project number
                if 'İŞ EMRİ NO' in row_str or any('24057' in str(cell) for cell in row if pd.notna(cell)):
                    for cell in row:
                        if pd.notna(cell) and '24057' in str(cell):
                            project_info['is_emri_no'] = str(cell)
                            break
                
                # Extract other information
                if 'PROJE ADI' in row_str:
                    next_cells = [str(cell) for cell in row if pd.notna(cell) and 'PROJE ADI' not in str(cell)]
                    if next_cells:
                        project_info['proje_adi'] = next_cells[0]
                
                if 'REVIZYON NO' in row_str:
                    for cell in row:
                        if pd.notna(cell) and str(cell).startswith('R'):
                            project_info['revizyon_no'] = str(cell)
                            break
                            
        except Exception as e:
            print(f"Warning: Could not extract project info: {e}")
            
        return project_info

    def find_latest_excel_files(self, directory='.', pattern='*.xlsx'):
        """Find the two most recent WSCAD Excel files in a directory"""
        try:
            # Get all Excel files (both .xlsx and .xls)
            xlsx_files = glob.glob(os.path.join(directory, '*.xlsx'))
            xls_files = glob.glob(os.path.join(directory, '*.xls'))
            all_files = xlsx_files + xls_files

            # Filter for WSCAD Excel files
            wscad_files = []
            for f in all_files:
                if os.path.isfile(f) and not os.path.basename(f).startswith('.'):
                    try:
                        if self.is_wscad_excel(f):
                            wscad_files.append(f)
                    except Exception as e:
                        print(f"WSCAD format check error: {e}")
                        continue

            if len(wscad_files) < 2:
                raise ValueError(f"At least two WSCAD Excel files required in directory: {directory}")

            # Sort by modification time (newest first)
            wscad_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)

            return wscad_files[0], wscad_files[1]
        except Exception as e:
            raise Exception(f"Error finding Excel files: {e}")

    def compare_excel_files(self, filepath1, filepath2, username=None):
        """Compare two WSCAD Excel files with focus on BOM-specific changes"""
        try:
            # Validate file paths
            if not filepath1 or not filepath2:
                raise ValueError("File paths cannot be empty")

            # Check if files exist
            if not os.path.exists(filepath1):
                raise FileNotFoundError(f"First file not found: {filepath1}")
            if not os.path.exists(filepath2):
                raise FileNotFoundError(f"Second file not found: {filepath2}")

            # Load Excel files with proper header detection
            df1 = pd.read_excel(filepath1, sheet_name=self.default_sheet_name, header=self.header_row-1)
            df2 = pd.read_excel(filepath2, sheet_name=self.default_sheet_name, header=self.header_row-1)
            
            # Clean column names
            df1.columns = [str(col).strip().replace('\r\n', '\n') for col in df1.columns]
            df2.columns = [str(col).strip().replace('\r\n', '\n') for col in df2.columns]

            # Remove empty rows (where POZ NO is NaN or empty)
            df1 = df1.dropna(subset=['POZ NO']).reset_index(drop=True)
            df2 = df2.dropna(subset=['POZ NO']).reset_index(drop=True)
            
            comparison_results = []

            # 1. Structure comparison
            structure_diff = self._compare_structure(df1, df2)
            comparison_results.extend(structure_diff)

            # 2. BOM-specific comparisons
            bom_diff = self._compare_bom_data(df1, df2, username)
            comparison_results.extend(bom_diff)

            # Sort results by importance: structure changes first, then by POZ NO
            comparison_results.sort(key=lambda x: (
                0 if x['type'] == 'structure' else 1,
                x.get('poz_no', 999999),  # Sort by POZ NO if available
                0 if x.get('change_type') == 'removed' else (1 if x.get('change_type') == 'added' else 2)
            ))

            return comparison_results
            
        except Exception as e:
            raise Exception(f"Error comparing WSCAD Excel files: {e}")

    def _compare_structure(self, df1, df2):
        """Compare structural differences between DataFrames"""
        structure_diff = []

        # Compare row counts
        if len(df1) != len(df2):
            structure_diff.append({
                'type': 'structure',
                'element': 'row_count',
                'value1': len(df1),
                'value2': len(df2),
                'diff': abs(len(df1) - len(df2)),
                'description': f"BOM item count changed: {len(df1)} → {len(df2)}",
                'severity': 'high'
            })

        # Compare column structure
        cols1 = set(df1.columns)
        cols2 = set(df2.columns)

        added_cols = cols2 - cols1
        if added_cols:
            structure_diff.append({
                'type': 'structure',
                'element': 'added_columns',
                'value1': '',
                'value2': ', '.join(added_cols),
                'diff': len(added_cols),
                'description': f"New columns added: {', '.join(added_cols)}",
                'severity': 'medium'
            })

        removed_cols = cols1 - cols2
        if removed_cols:
            structure_diff.append({
                'type': 'structure',
                'element': 'removed_columns',
                'value1': ', '.join(removed_cols),
                'value2': '',
                'diff': len(removed_cols),
                'description': f"Columns removed: {', '.join(removed_cols)}",
                'severity': 'high'
            })

        return structure_diff

    def _compare_bom_data(self, df1, df2, username=None):
        """Compare BOM data with focus on critical fields"""
        bom_diff = []
        common_cols = set(df1.columns).intersection(set(df2.columns))
        
        # Create dictionaries for faster lookup using POZ NO as key
        df1_dict = {}
        df2_dict = {}
        
        try:
            # Build lookup dictionaries
            for idx, row in df1.iterrows():
                poz_no = str(row.get('POZ NO', idx)).strip()
                df1_dict[poz_no] = row
                
            for idx, row in df2.iterrows():
                poz_no = str(row.get('POZ NO', idx)).strip()
                df2_dict[poz_no] = row
        except Exception as e:
            print(f"Warning: Error building lookup dictionaries: {e}")
            # Fallback to index-based comparison
            return self._compare_by_index(df1, df2, common_cols, username)

        all_poz_nos = set(df1_dict.keys()).union(set(df2_dict.keys()))

        for poz_no in sorted(all_poz_nos, key=lambda x: int(x) if x.isdigit() else 999999):
            row1 = df1_dict.get(poz_no)
            row2 = df2_dict.get(poz_no)

            if row1 is None:
                # New item added
                bom_diff.append({
                    'type': 'bom_item',
                    'poz_no': poz_no,
                    'column': 'ENTIRE_ROW',
                    'value1': '',
                    'value2': f"New item: {row2.get('PARCA ADI', 'Unknown')}",
                    'change_type': 'added',
                    'severity': 'high',
                    'description': f"POZ {poz_no}: New BOM item added",
                    'modified_by': username or 'System',
                    'modified_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                continue

            if row2 is None:
                # Item removed
                bom_diff.append({
                    'type': 'bom_item',
                    'poz_no': poz_no,
                    'column': 'ENTIRE_ROW',
                    'value1': f"Removed item: {row1.get('PARCA ADI', 'Unknown')}",
                    'value2': '',
                    'change_type': 'removed',
                    'severity': 'high',
                    'description': f"POZ {poz_no}: BOM item removed",
                    'modified_by': username or 'System',
                    'modified_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                continue

            # Compare each column for this POZ NO
            for col in common_cols:
                if col not in row1.index or col not in row2.index:
                    continue
                    
                val1 = str(row1[col]).strip() if pd.notna(row1[col]) else ''
                val2 = str(row2[col]).strip() if pd.notna(row2[col]) else ''

                if val1 != val2:
                    # Determine severity based on column importance
                    severity = 'high' if col in self.critical_columns else 'medium'
                    
                    # Special handling for quantity changes
                    change_type = 'modified'
                    if col in ['TOPLAM\nADET', 'BİRİM\nADET']:
                        try:
                            old_qty = float(val1) if val1 else 0
                            new_qty = float(val2) if val2 else 0
                            if new_qty > old_qty:
                                change_type = 'quantity_increased'
                            elif new_qty < old_qty:
                                change_type = 'quantity_decreased'
                            if new_qty == 0:
                                change_type = 'quantity_zeroed'
                                severity = 'high'
                        except ValueError:
                            pass

                    bom_diff.append({
                        'type': 'bom_field',
                        'poz_no': poz_no,
                        'column': col,
                        'value1': val1,
                        'value2': val2,
                        'change_type': change_type,
                        'severity': severity,
                        'description': f"POZ {poz_no} - {col}: '{val1}' → '{val2}'",
                        'modified_by': username or 'System',
                        'modified_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })

        return bom_diff

    def _compare_by_index(self, df1, df2, common_cols, username=None):
        """Fallback method: compare by row index when POZ NO lookup fails"""
        bom_diff = []
        max_rows = max(len(df1), len(df2))
        
        for idx in range(max_rows):
            row1 = df1.iloc[idx] if idx < len(df1) else None
            row2 = df2.iloc[idx] if idx < len(df2) else None
            
            poz_no = None
            if row1 is not None and 'POZ NO' in row1.index:
                poz_no = str(row1['POZ NO']) if pd.notna(row1['POZ NO']) else str(idx)
            elif row2 is not None and 'POZ NO' in row2.index:
                poz_no = str(row2['POZ NO']) if pd.notna(row2['POZ NO']) else str(idx)
            else:
                poz_no = str(idx)

            if row1 is None:
                # New row added
                bom_diff.append({
                    'type': 'bom_item',
                    'poz_no': poz_no,
                    'column': 'ENTIRE_ROW',
                    'value1': '',
                    'value2': 'New row added',
                    'change_type': 'added',
                    'severity': 'high'
                })
                continue

            if row2 is None:
                # Row removed
                bom_diff.append({
                    'type': 'bom_item',
                    'poz_no': poz_no,
                    'column': 'ENTIRE_ROW',
                    'value1': 'Row removed',
                    'value2': '',
                    'change_type': 'removed',
                    'severity': 'high'
                })
                continue

            # Compare columns
            for col in common_cols:
                if col not in row1.index or col not in row2.index:
                    continue
                    
                val1 = str(row1[col]).strip() if pd.notna(row1[col]) else ''
                val2 = str(row2[col]).strip() if pd.notna(row2[col]) else ''

                if val1 != val2:
                    bom_diff.append({
                        'type': 'bom_field',
                        'poz_no': poz_no,
                        'column': col,
                        'value1': val1,
                        'value2': val2,
                        'change_type': 'modified',
                        'severity': 'high' if col in self.critical_columns else 'medium',
                        'modified_by': username or 'System',
                        'modified_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })

        return bom_diff

    def generate_comparison_report(self, comparison_results):
        """Generate an enhanced Excel report for WSCAD BOM comparison"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BOM Karşılaştırma Raporu"

            # Enhanced header
            ws.cell(row=1, column=1, value="WSCAD BOM Karşılaştırma Raporu")
            ws.merge_cells('A1:J1')
            ws['A1'].font = Font(bold=True, size=16, color="0000FF")
            ws['A1'].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            # Timestamp
            ws.cell(row=2, column=1, value=f"Oluşturulma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
            ws['A2'].font = Font(italic=True)
            ws.merge_cells('A2:J2')

            # Headers specific to WSCAD BOM
            headers = ["Tür", "POZ NO", "Sütun", "Eski Değer", "Yeni Değer", "Değişiklik", "Önem", "Değiştiren", "Tarih", "Açıklama"]
            row_offset = 4

            # Header styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # Add data with WSCAD-specific formatting
            for row_idx, diff in enumerate(comparison_results, row_offset):
                # Type column
                type_cell = ws.cell(row=row_idx, column=1)
                if diff.get('type') == 'structure':
                    type_cell.value = "Yapı"
                elif diff.get('type') == 'bom_item':
                    type_cell.value = "BOM Kalemi"
                elif diff.get('type') == 'bom_field':
                    type_cell.value = "Alan"
                type_cell.font = Font(bold=True)

                # POZ NO
                ws.cell(row=row_idx, column=2, value=diff.get('poz_no', ''))

                # Column name
                ws.cell(row=row_idx, column=3, value=diff.get('column', ''))

                # Values
                orig_cell = ws.cell(row=row_idx, column=4, value=diff.get('value1', ''))
                new_cell = ws.cell(row=row_idx, column=5, value=diff.get('value2', ''))

                # Change type with BOM-specific colors
                change_type = diff.get('change_type', '')
                change_cell = ws.cell(row=row_idx, column=6)
                
                if change_type == 'added':
                    new_cell.fill = self.added_fill
                    change_cell.value = "Eklendi"
                elif change_type == 'removed':
                    orig_cell.fill = self.removed_fill
                    change_cell.value = "Silindi"
                elif change_type == 'quantity_increased':
                    new_cell.fill = self.quantity_increase_fill
                    change_cell.value = "Miktar Arttı"
                elif change_type == 'quantity_decreased':
                    new_cell.fill = self.quantity_decrease_fill
                    change_cell.value = "Miktar Azaldı"
                elif change_type == 'quantity_zeroed':
                    new_cell.fill = self.zero_quantity_fill
                    change_cell.value = "Miktar Sıfırlandı"
                else:
                    orig_cell.fill = self.changed_fill
                    new_cell.fill = self.changed_fill
                    change_cell.value = "Değiştirildi"

                # Severity
                severity_cell = ws.cell(row=row_idx, column=7, value=diff.get('severity', 'medium'))
                if diff.get('severity') == 'high':
                    severity_cell.font = Font(bold=True, color="FF0000")
                elif diff.get('severity') == 'medium':
                    severity_cell.font = Font(color="FF8C00")

                # Modified by and date
                ws.cell(row=row_idx, column=8, value=diff.get('modified_by', 'System'))
                ws.cell(row=row_idx, column=9, value=diff.get('modified_date', ''))

                # Description
                ws.cell(row=row_idx, column=10, value=diff.get('description', ''))

            # Column widths optimization for WSCAD data
            column_widths = {
                1: 12,  # Tür
                2: 10,  # POZ NO
                3: 20,  # Sütun
                4: 35,  # Eski Değer
                5: 35,  # Yeni Değer
                6: 18,  # Değişiklik
                7: 10,  # Önem
                8: 15,  # Değiştiren
                9: 20,  # Tarih
                10: 50  # Açıklama
            }

            for col, width in column_widths.items():
                column_letter = openpyxl.utils.get_column_letter(col)
                ws.column_dimensions[column_letter].width = width

            # Apply borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in range(row_offset, len(comparison_results) + row_offset):
                for col in range(1, 11):
                    ws.cell(row=row, column=col).border = thin_border

            # Create enhanced summary sheet
            summary_ws = wb.create_sheet(title="BOM Özeti")
            self._create_bom_summary(summary_ws, comparison_results)

            # Create change statistics sheet
            stats_ws = wb.create_sheet(title="İstatistikler")
            self._create_statistics_sheet(stats_ws, comparison_results)

            # Save to byte stream
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return output
        except Exception as e:
            raise Exception(f"Error generating WSCAD BOM comparison report: {e}")

    def _create_bom_summary(self, ws, comparison_results):
        """Create BOM-specific summary sheet"""
        ws.cell(row=1, column=1, value="BOM Karşılaştırma Özeti").font = Font(bold=True, size=14)
        
        # Calculate BOM-specific statistics
        structure_changes = sum(1 for diff in comparison_results if diff.get('type') == 'structure')
        bom_item_changes = sum(1 for diff in comparison_results if diff.get('type') == 'bom_item')
        field_changes = sum(1 for diff in comparison_results if diff.get('type') == 'bom_field')
        
        added_items = sum(1 for diff in comparison_results if diff.get('change_type') == 'added')
        removed_items = sum(1 for diff in comparison_results if diff.get('change_type') == 'removed')
        quantity_increases = sum(1 for diff in comparison_results if diff.get('change_type') == 'quantity_increased')
        quantity_decreases = sum(1 for diff in comparison_results if diff.get('change_type') == 'quantity_decreased')
        quantity_zeroed = sum(1 for diff in comparison_results if diff.get('change_type') == 'quantity_zeroed')
        
        high_severity = sum(1 for diff in comparison_results if diff.get('severity') == 'high')
        medium_severity = sum(1 for diff in comparison_results if diff.get('severity') == 'medium')

        # Summary data
        summary_data = [
            ("", ""),  # Empty row
            ("GENEL İSTATİSTİKLER", ""),
            ("Yapısal Değişiklikler", structure_changes),
            ("BOM Kalem Değişiklikleri", bom_item_changes),
            ("Alan Değişiklikleri", field_changes),
            ("Toplam Değişiklik", len(comparison_results)),
            ("", ""),
            ("KALEM DEĞİŞİKLİKLERİ", ""),
            ("Eklenen Kalemler", added_items),
            ("Çıkarılan Kalemler", removed_items),
            ("", ""),
            ("MİKTAR DEĞİŞİKLİKLERİ", ""),
            ("Miktar Artışları", quantity_increases),
            ("Miktar Azalışları", quantity_decreases),
            ("Sıfırlanan Miktarlar", quantity_zeroed),
            ("", ""),
            ("ÖNEMLİLİK DAĞILIMI", ""),
            ("Yüksek Önemde", high_severity),
            ("Orta Önemde", medium_severity)
        ]

        # Headers
        ws.cell(row=3, column=1, value="Kategori").font = Font(bold=True)
        ws.cell(row=3, column=2, value="Sayı").font = Font(bold=True)

        for idx, (category, count) in enumerate(summary_data, 4):
            cell_a = ws.cell(row=idx, column=1, value=category)
            cell_b = ws.cell(row=idx, column=2, value=count)
            
            if category and not str(count).isdigit():  # Section headers
                cell_a.font = Font(bold=True, color="0000FF")
            elif str(count).isdigit() and int(count) > 0:  # Non-zero counts
                cell_b.font = Font(bold=True, color="FF6600")

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

        # Add color legend
        legend_row = len(summary_data) + 8
        ws.cell(row=legend_row, column=1, value="Renk Açıklamaları:").font = Font(bold=True)
        
        legend_items = [
            (self.added_fill, "Yeni Eklenen Kalemler"),
            (self.removed_fill, "Çıkarılan Kalemler"),
            (self.quantity_increase_fill, "Miktar Artışları"),
            (self.quantity_decrease_fill, "Miktar Azalışları"),
            (self.zero_quantity_fill, "Sıfırlanan Miktarlar"),
            (self.changed_fill, "Genel Değişiklikler")
        ]
        
        for i, (color_fill, description) in enumerate(legend_items):
            current_row = legend_row + i + 1
            cell = ws.cell(row=current_row, column=1, value="")
            cell.fill = color_fill
            ws.cell(row=current_row, column=2, value=description)

    def _create_statistics_sheet(self, ws, comparison_results):
        """Create detailed statistics sheet"""
        ws.cell(row=1, column=1, value="Detaylı İstatistikler").font = Font(bold=True, size=14)
        
        # Group changes by POZ NO
        poz_changes = {}
        column_changes = {}
        
        for diff in comparison_results:
            poz_no = diff.get('poz_no', 'Unknown')
            column = diff.get('column', 'Unknown')
            
            if poz_no not in poz_changes:
                poz_changes[poz_no] = 0
            poz_changes[poz_no] += 1
            
            if column not in column_changes:
                column_changes[column] = 0
            column_changes[column] += 1

        # Most changed POZ NOs
        ws.cell(row=3, column=1, value="En Çok Değişen POZ Numaraları").font = Font(bold=True)
        ws.cell(row=4, column=1, value="POZ NO").font = Font(bold=True)
        ws.cell(row=4, column=2, value="Değişiklik Sayısı").font = Font(bold=True)
        
        sorted_poz = sorted(poz_changes.items(), key=lambda x: x[1], reverse=True)[:10]
        for idx, (poz_no, count) in enumerate(sorted_poz, 5):
            ws.cell(row=idx, column=1, value=poz_no)
            ws.cell(row=idx, column=2, value=count)

        # Most changed columns
        ws.cell(row=3, column=4, value="En Çok Değişen Sütunlar").font = Font(bold=True)
        ws.cell(row=4, column=4, value="Sütun").font = Font(bold=True)
        ws.cell(row=4, column=5, value="Değişiklik Sayısı").font = Font(bold=True)
        
        sorted_cols = sorted(column_changes.items(), key=lambda x: x[1], reverse=True)[:10]
        for idx, (column, count) in enumerate(sorted_cols, 5):
            ws.cell(row=idx, column=4, value=column)
            ws.cell(row=idx, column=5, value=count)

        # Column widths
        for col in ['A', 'B', 'D', 'E']:
            ws.column_dimensions[col].width = 20

    def auto_compare_latest_files(self, directory='.', username=None):
        """Automatically find and compare the two most recent WSCAD Excel files"""
        try:
            print("WSCAD BOM otomatik karşılaştırma başlatılıyor...")
            
            # Find Excel files in directory
            excel_files = self.list_excel_files(directory)
            if len(excel_files) < 2:
                raise ValueError(f"En az iki WSCAD Excel dosyası gerekli. Dizinde {len(excel_files)} dosya bulundu.")

            print(f"Bulunan WSCAD Excel dosyaları: {len(excel_files)}")

            # Get the two most recent files
            file1 = excel_files[0]['filepath']
            file2 = excel_files[1]['filepath']

            print(f"Karşılaştırılıyor:\n1. {os.path.basename(file1)}\n2. {os.path.basename(file2)}")

            # Process both files
            info1 = self.process_file(file1)
            info2 = self.process_file(file2)

            # Compare the files
            comparison_data = self.compare_excel_files(file1, file2, username)

            # Create comparison result structure
            comparison_result = {
                'file1': info1,
                'file2': info2,
                'comparison_data': comparison_data,
                'comparison_count': len(comparison_data),
                'comparison_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'modified_by': username if username else 'System'
            }

            # Generate report
            report_data = self.generate_comparison_report(comparison_data)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_filename = f"wscad_bom_comparison_{timestamp}.xlsx"

            # Save the report
            with open(report_filename, "wb") as f:
                f.write(report_data.getvalue())

            return {
                'file1': info1,
                'file2': info2,
                'comparison_count': len(comparison_data),
                'comparison_data': comparison_data,
                'report_file': report_filename
            }
        except Exception as e:
            raise Exception(f"WSCAD BOM otomatik karşılaştırma hatası: {e}")

    def compare_specific_files(self, filepath1, filepath2, username=None):
        """Compare two specific WSCAD Excel files"""
        try:
            if not os.path.exists(filepath1):
                raise FileNotFoundError(f"İlk dosya bulunamadı: {filepath1}")
            if not os.path.exists(filepath2):
                raise FileNotFoundError(f"İkinci dosya bulunamadı: {filepath2}")

            print(f"WSCAD BOM dosyaları karşılaştırılıyor:\n1. {filepath1}\n2. {filepath2}")

            # Process both files
            info1 = self.process_file(filepath1)
            info2 = self.process_file(filepath2)

            # Compare the files
            comparison_results = self.compare_excel_files(filepath1, filepath2, username)

            # Generate report
            report_data = self.generate_comparison_report(comparison_results)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_filename = f"wscad_bom_comparison_{timestamp}.xlsx"

            # Save the report
            with open(report_filename, "wb") as f:
                f.write(report_data.getvalue())

            return {
                'file1': info1,
                'file2': info2,
                'comparison_count': len(comparison_results),
                'report_file': report_filename
            }
        except Exception as e:
            raise Exception(f"WSCAD BOM dosya karşılaştırma hatası: {e}")

    def prepare_for_export(self, filepath):
        """Prepare WSCAD Excel data for export to ERP system"""
        try:
            if not os.path.exists(filepath):
                raise FileNotFoundError(f"File not found: {filepath}")

            # Load the Excel file with proper header detection
            df = pd.read_excel(filepath, sheet_name=self.default_sheet_name, header=self.header_row-1)
            
            # Clean column names
            df.columns = [str(col).strip().replace('\r\n', '\n') for col in df.columns]
            
            # Remove empty rows
            df = df.dropna(subset=['POZ NO']).reset_index(drop=True)

            # Get file and project information
            file_info = self.process_file(filepath)

            # Convert DataFrame to ERP-friendly format
            data_rows = []
            for _, row in df.iterrows():
                cleaned_row = {}
                for col, value in row.items():
                    if pd.isna(value):
                        cleaned_row[str(col)] = None
                    else:
                        # Special handling for numeric fields
                        if col in ['POZ NO', 'TOPLAM\nADET', 'BİRİM\nADET', 'AGIRLIK']:
                            try:
                                cleaned_row[str(col)] = float(value) if value != '' else 0
                            except (ValueError, TypeError):
                                cleaned_row[str(col)] = str(value)
                        else:
                            cleaned_row[str(col)] = str(value)

                data_rows.append(cleaned_row)

            # Create ERP export structure
            export_data = {
                'file_info': file_info,
                'project_info': file_info.get('project_info', {}),
                'bom_data': data_rows,
                'export_timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }

            return export_data
        except Exception as e:
            raise Exception(f"WSCAD Excel ERP export hazırlama hatası: {e}")

    def list_excel_files(self, directory='.'):
        """List all WSCAD Excel files in the specified directory"""
        try:
            # Get all Excel files
            xlsx_files = glob.glob(os.path.join(directory, '*.xlsx'))
            xls_files = glob.glob(os.path.join(directory, '*.xls'))
            excel_files = xlsx_files + xls_files

            # Filter for WSCAD files and valid files
            wscad_files = []
            for file in excel_files:
                if os.path.isfile(file) and not os.path.basename(file).startswith('.'):
                    try:
                        if self.is_wscad_excel(file):
                            wscad_files.append(file)
                    except Exception:
                        continue

            # Sort by modification time (newest first)
            wscad_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)

            # Create detailed file list
            file_list = []
            for file in wscad_files:
                try:
                    file_info = {
                        'filepath': file,
                        'filename': os.path.basename(file),
                        'size_kb': round(os.path.getsize(file) / 1024, 2),
                        'modified': datetime.fromtimestamp(os.path.getmtime(file)).strftime('%Y-%m-%d %H:%M:%S'),
                        'is_wscad': True
                    }
                    
                    # Try to extract project info quickly
                    try:
                        header_df = pd.read_excel(file, sheet_name=self.default_sheet_name, header=None, nrows=6)
                        project_info = self._extract_project_info(header_df)
                        file_info['project_info'] = project_info
                    except Exception:
                        file_info['project_info'] = {}
                    
                    file_list.append(file_info)
                except Exception as e:
                    print(f"Warning: Could not process file {file}: {e}")
                    continue

            return file_list
        except Exception as e:
            raise Exception(f"WSCAD Excel dosyalarını listelerken hata: {e}")


# Example usage for WSCAD BOM files
if __name__ == "__main__":
    try:
        processor = ExcelProcessor()

        print("WSCAD BOM Excel İşlemcisi")
        print("1. Son iki WSCAD BOM dosyasını otomatik karşılaştır")
        print("2. Belirli iki WSCAD BOM dosyasını karşılaştır")
        print("3. Mevcut WSCAD BOM dosyalarını listele")
        print("4. WSCAD BOM dosyasını ERP'ye hazırla")
        print("5. Çıkış")

        choice = input("İşlem seçin (1-5): ")

        if choice == "1":
            directory = input("Dizin yolu (Enter = mevcut dizin): ") or '.'
            username = input("Kullanıcı adınız: ")
            result = processor.auto_compare_latest_files(directory, username)

            print(f"\nWO Karşılaştırma tamamlandı!")
            print(f"İlk dosya: {result['file1']['filename']}")
            print(f"İkinci dosya: {result['file2']['filename']}")
            print(f"Toplam {result['comparison_count']} BOM değişikliği tespit edildi")
            print(f"Rapor dosyası: {result['report_file']}")

        elif choice == "2":
            files = processor.list_excel_files()
            
            if len(files) < 2:
                print("Karşılaştırma için en az iki WSCAD BOM dosyası gerekli!")
            else:
                print("\nMevcut WSCAD BOM Dosyaları:")
                for i, file in enumerate(files, 1):
                    project_no = file.get('project_info', {}).get('is_emri_no', 'Bilinmiyor')
                    print(f"{i}. {file['filename']} - Proje: {project_no} - {file['modified']}")

                try:
                    file1_idx = int(input("\nBirinci dosya numarası: ")) - 1
                    file2_idx = int(input("İkinci dosya numarası: ")) - 1

                    if 0 <= file1_idx < len(files) and 0 <= file2_idx < len(files):
                        username = input("Kullanıcı adınız: ")
                        result = processor.compare_specific_files(
                            files[file1_idx]['filepath'], 
                            files[file2_idx]['filepath'], 
                            username
                        )

                        print(f"\nWO Karşılaştırma tamamlandı!")
                        print(f"Toplam {result['comparison_count']} BOM değişikliği tespit edildi")
                        print(f"Rapor dosyası: {result['report_file']}")
                    else:
                        print("Geçersiz dosya numarası!")
                except ValueError:
                    print("Lütfen geçerli bir sayı girin!")

        elif choice == "3":
            directory = input("Dizin yolu (Enter = mevcut dizin): ") or '.'
            files = processor.list_excel_files(directory)

            if not files:
                print(f"Belirtilen dizinde WSCAD BOM dosyası bulunamadı!")
            else:
                print(f"\nMevcut WSCAD BOM Dosyaları ({len(files)}):")
                for i, file in enumerate(files, 1):
                    project_info = file.get('project_info', {})
                    project_no = project_info.get('is_emri_no', 'N/A')
                    project_name = project_info.get('proje_adi', 'N/A')
                    print(f"{i}. {file['filename']}")
                    print(f"   Proje No: {project_no}")
                    print(f"   Proje Adı: {project_name}")
                    print(f"   Boyut: {file['size_kb']} KB - Değiştirilme: {file['modified']}")
                    print()

        elif choice == "4":
            files = processor.list_excel_files()
            
            if not files:
                print("ERP'ye aktarılacak WSCAD BOM dosyası bulunamadı!")
            else:
                print("\nMevcut WSCAD BOM Dosyaları:")
                for i, file in enumerate(files, 1):
                    project_no = file.get('project_info', {}).get('is_emri_no', 'Bilinmiyor')
                    print(f"{i}. {file['filename']} - Proje: {project_no}")

                try:
                    file_idx = int(input("\nERP'ye aktarılacak dosya numarası: ")) - 1
                    if 0 <= file_idx < len(files):
                        export_data = processor.prepare_for_export(files[file_idx]['filepath'])
                        
                        # Save export data as JSON
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        export_filename = f"wscad_bom_erp_export_{timestamp}.json"
                        
                        with open(export_filename, 'w', encoding='utf-8') as f:
                            json.dump(export_data, f, ensure_ascii=False, indent=2, default=str)
                        
                        print(f"\nERP export dosyası hazırlandı: {export_filename}")
                        print(f"Toplam BOM kalemi: {len(export_data['bom_data'])}")
                    else:
                        print("Geçersiz dosya numarası!")
                except ValueError:
                    print("Lütfen geçerli bir sayı girin!")

        elif choice == "5":
            print("WSCAD BOM İşlemcisi sonlandırılıyor...")

        else:
            print("Geçersiz seçim!")

    except Exception as e:
        print(f"Hata: {e}")